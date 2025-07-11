import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url = "https://www.malware-traffic-analysis.net/2023/index.html"
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36'}

# 웹 요청
r = requests.get(url, headers=headers, verify=False)
soup = BeautifulSoup(r.text, 'html.parser')

# 태그 추출
tags = soup.select("#main_content > div.content > ul > li > a.main_menu")

# 엑셀 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "Malware URL"

# 헤더 추가
ws.append(["설명", "링크"])

# 데이터 추가
for tag in tags:
    title = tag.text.strip()
    link = "https://www.malware-traffic-analysis.net/2023/" + tag.get("href")
    ws.append([title, link])
    print(title, link)

# 엑셀 저장
wb.save("malware.xlsx")

"""
wb = Workbook()
ws = wb.active
ws['A1'] = "설명"
ws['B1'] = "URL 링크"

i = 2
for tag in tags:
    ws.cell(row=i, column=1, value=tag.text)
    ws.cell(row=i, column=2, value=f"https://www.malware-traffic-analysis.net/2023/{tag.get('href')}")
    i = i + 1

wb.save("malwares.xlsx")
"""