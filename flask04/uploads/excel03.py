from openpyxl import load_workbook

book = load_workbook("excel_data2.xlsx", data_only=True)
ws = book.active

for row in ws["A1":"E7"]:
    result = []
    for cell in row:
        result.append(cell.value)
    print(result)

for row in ws["A1":"E7"]:
    values = [cell.value for cell in row]
